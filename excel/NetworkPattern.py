import re
import openpyxl
import sys
import ntpath

from openpyxl.descriptors.base import String
sys.path.append('/.../openess/excel')
import Cell

##################################
class NetworkPattern:

    __sheetNames = ["Arguments", "Result"]
    __cellName = "A2"
    __cellArgs = "B1"
    __cellResults = "A1"
    __argsPrefix = "Arg"
    __resultPrefix = "$"

    __wbObj = None
    
    ##################################
    def __init__(self, path):
        self.__wbObj = openpyxl.load_workbook(path)
        self.__checkFile((ntpath.basename(path)).split('.')[0])
        # print(self.getArgsNumber())
        # print(self.getResultNumber())
        # print(self.getFormulas())

    ##################################
    def __checkFile(self, fileName):

        i = 0
        for s in self.__sheetNames:
            if s != self.__wbObj.sheetnames[i]:
                raise Exception("Missing sheet: " + self.__sheetNames[i])
            i = i + 1

        sheet_obj = self.__wbObj[self.__sheetNames[0]]
        pos = Cell.cellPositionFromString(self.__cellName)
        cell_obj = sheet_obj.cell(pos[0], pos[1])
        if cell_obj.value == None:
            raise Exception("Function name cannot be empty")
        if fileName != cell_obj.value:
            raise Exception("Wrong function name: " + str(cell_obj.value))

    ##################################
    def getArgsNumber(self):
        sheet_obj = self.__wbObj[self.__sheetNames[0]]
        return NetworkPattern.__calcAmountOfArgs(sheet_obj, self.__cellArgs, self.__argsPrefix)

    ##################################  
    def getResultNumber(self):
        sheet_obj = self.__wbObj[self.__sheetNames[1]]
        return NetworkPattern.__calcAmountOfArgs(sheet_obj, self.__cellResults, self.__resultPrefix)

    ##################################  
    def getFormulas(self):
        formulas = []
        sheetObjResult = self.__wbObj[self.__sheetNames[1]]
        initColumnArgs = Cell.cellPositionFromString(self.__cellArgs)[1]
        posResult = Cell.cellPositionFromString(self.__cellResults)
        num = self.getResultNumber()

        i = 0
        while i < num:
            cell_obj = sheetObjResult.cell(posResult[0] + 1, posResult[1] + i)
            x = NetworkPattern.__convertFormula(cell_obj.value, initColumnArgs)
            formulas.append(x)
            i = i + 1
        
        return formulas

    ##################################  
    def __convertFormula(formula, initColumnArgs):
        ret = ""
        if isinstance(formula, str):
            i = 0
            ret = formula
            while i < len(ret):
                if ret[i] == '$':
                    pos = Cell.cellPositionFromString(ret[i+1:i+3])
                    arg = pos[1] - initColumnArgs
                    ret = ret[:i+1] + str(arg) + ret[i+3:]
                i = i + 1

        return ret

    ##################################       
    def __calcAmountOfArgs(sheet_obj, cellName, prefix):
        pos = Cell.cellPositionFromString(cellName)

        i = 0
        while True:
            cell_obj = sheet_obj.cell(pos[0], pos[1] + i)
            if NetworkPattern.__checkName(prefix, cell_obj.value, i):
                i = i + 1
            else:
                break
        return i

    ##################################  
    def __checkName(prefix, value, number):
        argsPrefixLen = len(prefix)
        if isinstance(value, str) and len(value) > argsPrefixLen:
            if value[:argsPrefixLen] == prefix:
                x = value[argsPrefixLen:]
                if x.isdigit:
                    if int(x) == number:
                        return True
        return False



networkPatternObj = NetworkPattern("C:\\pythonProject\\openess\\templates\\blokFunkcyjny.xlsx")
