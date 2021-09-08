import openpyxl
import sys
from openpyxl.styles import Border, Side, PatternFill, Font, GradientFill, Alignment
sys.path.append('/.../openess/excel')
import Cell
import NetworkPattern as pattern


##################################
class FunctionPattern:

    __sheetNames = ["Config", "Arguments", "Result"]
    __cellName = "A2"
    __cellNumber = "B2"
    __cellType = "C2"
    __cellNetworkName = "A2"
    __functionTypes = ["FC", "FB"]
    __networkPatternsPath = "C:\\pythonProject\\openess\\templates\\"

    __wbObj = None
    __wbPath = ""
    __amountOfNetworks = 0

    __networkPatterns = {}
    
    ##################################
    def __init__(self, path):
        self.__wbPath = path
        self.__wbObj = openpyxl.load_workbook(path)
        self.__amountOfNetworks = self.__calcAmountOfNetworks()
        self.__addNetworkPaterns()

        self.__checkFile()

    ##################################
    def __checkFile(self):

        i = 0
        for s in self.__sheetNames:
            if s != self.__wbObj.sheetnames[i]:
                raise Exception("Missing sheet: " + self.__sheetNames[i])
            i = i + 1

        sheet_obj = self.__wbObj[self.__sheetNames[0]]
        pos = Cell.cellPositionFromString(self.__cellName)
        cell_obj = sheet_obj.cell(pos[0], pos[1])
        if len(cell_obj.value) == 0:
            raise Exception("Wrong function name " + cell_obj.value)

        pos = Cell.cellPositionFromString(self.__cellNumber)
        cell_obj = sheet_obj.cell(pos[0], pos[1])

        if not isinstance(cell_obj.value, int):
            raise Exception("Wrong function number " + str(cell_obj.value))

        pos = Cell.cellPositionFromString(self.__cellType)
        cell_obj = sheet_obj.cell(pos[0], pos[1])

        functionTypeValid = False
        if len(cell_obj.value) > 0:
            for s in self.__functionTypes:
                if s == cell_obj.value:
                    functionTypeValid = True
                    break

        if not functionTypeValid:
            raise Exception("Wrong function type: " + str(cell_obj.value))

        sheet_obj = self.__wbObj[self.__sheetNames[1]]
        pos = Cell.cellPositionFromString(self.__cellNetworkName)
        i = 0

    ##################################
    def prepareFile(self):
        self.__colorArgs()
        self.__completeFormulas()
        return 0

    ##################################
    def __addNetworkPaterns(self):
        sheet_obj = self.__wbObj[self.__sheetNames[1]]
        pos = Cell.cellPositionFromString(self.__cellNetworkName)

        i = 0
        while True:
            ok = False
            cell_obj = sheet_obj.cell(pos[0] + i, pos[1])
            if isinstance(cell_obj.value, str):
                if len(cell_obj.value) > 0:
                    ok = True
                    if not (cell_obj.value + "" in self.__networkPatterns):
                        filePath = self.__networkPatternsPath + cell_obj.value + ".xlsx"
                        self.__networkPatterns[cell_obj.value] = pattern.NetworkPattern(filePath)
            i = i + 1
            if not ok:
                break

    ##################################      
    def __calcAmountOfNetworks(self):
        sheet_obj = self.__wbObj[self.__sheetNames[1]]
        pos = Cell.cellPositionFromString(self.__cellNetworkName)
        i = 0
        while True:
            ok = False
            cell_obj = sheet_obj.cell(pos[0] + i, pos[1])
            if isinstance(cell_obj.value, str):
                if len(cell_obj.value) > 0:
                    ok = True
            if not ok:
                break
            i = i + 1
        return i

    ##################################
    def __colorArgs(self):
        sheet_obj = self.__wbObj[self.__sheetNames[1]]
        pos = Cell.cellPositionFromString(self.__cellNetworkName)

        i = 0
        while i < self.__amountOfNetworks:
            cellObjName = sheet_obj.cell(pos[0] + i, pos[1])
            argsNumber = self.__networkPatterns[cellObjName.value].getArgsNumber()
            j = 0
            while j < argsNumber:
                cellObjName = sheet_obj.cell(pos[0] + i, pos[1] + 1 + j)
                cellObjName.fill = PatternFill("solid", fgColor="DDDDDD")

                j = j + 1 
            i = i + 1

    ##################################
    def __completeFormulas(self):
        FunctionPattern.__convertFormula('="zmienna"&Arguments!$1&Arguments!$0', [2,2])

        return 0

    ##################################
    def __convertFormula(formula, initPosArg):
        ret = ""

        if isinstance(formula, str):
            i = 0
            while i < len(formula):
                if formula[i] == "$":
                    j = i + 1
                    while j < len(formula) and formula[j].isdigit():
                        j = j + 1
                    print(formula[i+1:j])

 

                    posString = Cell.cellPositionFromNumbers( [i + initPosArg[0], int(formula[i+1:j]) + initPosArg[1]] )
                    print(posString)

                i = i + 1

        return ret

    

functionPatternObj = FunctionPattern("C:\\pythonProject\\openess\\templates\\testFc.xlsx")
functionPatternObj.prepareFile()

# functionPatternObj.showWbObj()

# functionPatternObj.showWbObj()

# To open the workbook
# workbook object is created
# wb_obj = openpyxl.load_workbook(path)
 
# Get workbook active sheet object
# from the active attribute
# sheet_obj = wb_obj.active
# sheet_obj = wb_obj["Arguments"]

# for s in sheet_obj:
#     print(s)

# print(wb_obj.sheetnames)

# Cell objects also have a row, column,
# and coordinate attributes that provide
# location information for the cell.
 
# Note: The first row or
# column integer is 1, not 0.
 
# Cell object is created by using
# sheet object's cell() method.
# cell_obj = sheet_obj.cell(row = 2, column = 1)
 
# Print value of cell object
# using the value attribute
# print(cell_obj.value)
